import fitz
import openpyxl
pdfdoc=fitz.open("D:\\My Internship Projects\\project01\\text extraction\\product1.pdf")
page1=pdfdoc.loadPage(0)
text=page1.getText()
lines=text.split("\n")
def spl(line):
	return [char for char in line]
def search(line):
	char=spl(line)
	for i in range(0,len(char)):
		if(char[i]==':'):
			if(i==len(char)-1):
				return 1
			else:
				return 2
	return 0
wb=openpyxl.load_workbook("D:\\My Internship Projects\\project01\\text extraction\\mytext3.xlsx")
sh=wb['Sheet1']
def pr(key,val,i):
	sh.cell(row=i,column=1,value=key)
	sh.cell(row=i,column=2,value=val)
def met(lines):
	j=sh.max_row-1
	for i in range(0,len(lines)):
		if(search(lines[i])==1):
			j=j+1
			pr(lines[i],lines[i+1],j)
		elif(search(lines[i])==2):
			j=j+1
			templines=lines[i].split(":")
			pr(templines[0],templines[1],j)

def allpage(pdfdoc):
	for k in range(0,pdfdoc.pageCount):
		lines=pdfdoc.loadPage(k).getText().split("\n")
		met(lines)
allpage(pdfdoc)
wb.save("D:\\My Internship Projects\\project01\\text extraction\\mytext3.xlsx")
